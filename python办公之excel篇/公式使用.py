from openpyxl import Workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import FORMULAE

wb = Workbook()
ws = wb.active

print(len(FORMULAE))

#判断求和公式是否在FORMULAE
print('SUM' in FORMULAE)

ws.append(['语文','数学','总成绩','平均成绩'])
ws.append([54,21])
ws.append([14,32])
ws.append([24,1])
ws.append([27,98])
ws['c2'] = '=SUM(A2:B2)'
ws['d2'] = '=AVERAGE(A2:B2)'

#翻译
ws['c3'] = Translator(formula='=SUM(A2:B2)',origin='c2').translate_formula('c3')
ws['c4'] = Translator(formula='=SUM(A2:B2)',origin='c2').translate_formula('c4')
ws['d3'] = Translator(formula='=AVERAGE(A2:B2)',origin='d2').translate_formula('d3')
ws['d4'] = Translator(formula='=AVERAGE(A2:B2)',origin='d2').translate_formula('d4')
#遍历
wb.save(r'C:\Users\肖阿强\Desktop\test.xlsx')

