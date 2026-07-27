from openpyxl import Workbook

wb = Workbook()
ws = wb.active
x = 1
for i in range(1,11):
    for j in range(1,11):
        ws.cell(i,j,x)
        x+=1

#合并单元格
ws.merge_cells('b2:d4')
#ws.merge_cells(2,4,2,4)  #效果与上述相同

#取消合并
ws.unmerge_cells('b2:d4')

#插入、删除、移动
ws.insert_cols(2,3)
ws.insert_rows(3,2)

ws.delete_cols(3,4)
ws.delete_rows(2,3)

ws.move_range('a2:b3',2,3)




wb.save(r'C:\Users\肖阿强\Desktop\test.xlsx')