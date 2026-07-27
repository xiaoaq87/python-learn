from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active
ws2 = wb.create_sheet('sheet2', 1)
ws3 = wb.create_sheet('sheet3', 2)
print(wb.sheetnames)
print(ws1.title)

wb.move_sheet('sheet2', 1)
print(wb.sheetnames)

ws4 = wb['sheet3']

del wb['sheet3']
print(ws3,ws4)

print(wb.sheetnames)
print(ws1.max_row)
print(ws1.max_column)

#复制工作表
ws5 = wb.copy_worksheet(ws1)
print(wb.sheetnames)


