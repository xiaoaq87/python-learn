from openpyxl import Workbook,load_workbook

#新建工作簿
#wb = Workbook() #注意Workbook 首字母是大写

#打开一个已建的工作簿
wb = load_workbook(r'C:\Users\肖阿强\Desktop\test.xlsx')


#默认激活一个工作表，这个值默认为0
ws = wb.active
#创建新的工作簿
#ws1 = wb.create_sheet('mysheet') 

#更改工作簿的名字
#ws1.title = '星辰大海'  

#工作表的名字
print(ws.title)

#查看工作簿的所有工作表的名字
print(wb.sheetnames) 

#保存工作簿
#wb.save(r'C:\Users\肖阿强\Desktop\test.xlsx')