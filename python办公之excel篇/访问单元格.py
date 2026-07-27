from openpyxl import Workbook

wb = Workbook()
ws = wb.active
#访问、修改
# ws['a6'] = '太阳'
# print(ws['a6'].value)

# cell = ws.cell(3,4)
# cell.value = '月亮'
# print(cell.value)

# print(cell.coordinate)   #反映坐标，列用字母反映
# print(cell.row)
# print(cell.column)        #用数字反映列
# print(cell.col_idx)        #用数字反映列，与上面相同
# print(cell.column_letter)  #用字母反映列
# print(cell.row)

x = 1
for i in range(1,10):
    for j in range(1,10):
        ws.cell(i,j,x)
        x+=1

print(ws['a:c'])
print(ws[1:5])
print(ws['a1:c4'])
print(type(ws['a1:c4']))  #返回的一个元组套元组
print(ws.iter_cols())  #遍历所列
print(ws.iter_rows())  #遍历所行

wb.save(r'C:\Users\肖阿强\Desktop\test.xlsx')